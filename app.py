from flask import Flask, render_template, request, redirect, session
from pymongo import MongoClient
from datetime import datetime
from zoneinfo import ZoneInfo
import os
import qrcode
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
    HRFlowable
)
from dotenv import load_dotenv

load_dotenv()
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from flask import Response
from bson import ObjectId
from flask import jsonify
import base64
# from reportlab.pdfbase import pdfmetrics
# from reportlab.pdfbase.ttfonts import TTFont
# from reportlab.lib.styles import ParagraphStyle
# import arabic_reshaper
# from bidi.algorithm import get_display

from openpyxl import Workbook
from io import BytesIO
from flask import Response
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment



app = Flask(__name__)
app.secret_key = "starcare_secret"
app.config['SESSION_COOKIE_SAMESITE'] = "Lax"
# font_path = os.path.join(app.root_path, "static/fonts/NotoNaskhArabic-Regular.ttf")

# pdfmetrics.registerFont(
#     TTFont('Arabic', font_path)
# )

# def fix_arabic(text):
#     if not text:
#         return text
#     return get_display(arabic_reshaper.reshape(str(text)))

def safe_int(v):
    try:
        return int(v)
    except:
        return 0


def safe_float(v):
    try:
        return float(v)
    except:
        return 0
# ---------------- MONGO (Atlas) ----------------
MONGO_URI = os.environ.get("MONGO_URI")

if not MONGO_URI:
    raise Exception("MONGO_URI is missing in environment variables")

client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
db = client["starcare_feedback"]
collection = db["feedback"]
users_collection = db["users"]
alerts_collection = db.alerts

# db.users.insert_one({
#     "username": "admin",
#     "password": "1234",
#     "role": "admin"
# })

# db.users.insert_one({
#     "username": "it_head",
#     "password": "1234",
#     "role": "it"
# })
# ---------------- BASE URL ----------------
BASE_URL = "https://starcare-feedback-1.onrender.com"

# ---------------- BRANCHES ----------------
# branches = ["alhail", "mabella", "alamerat"]

users = {
    "admin": {"password": "1234", "role": "admin"},
    "it_head": {"password": "1111", "role": "it"},
    "alhail_head": {"password": "2222", "role": "alhail"},
    "mabella_head": {"password": "3333", "role": "mabella"},
}
# ---------------- LOCATIONS ----------------
locations = [
    "consultation101","consultation102","consultation103","consultation104",
    "cystoscopy110","urodynamic109","waiting_area",
    "laboratory107","staff","ultrasound106","xray108",
    "triage105","nurse","department","doctor","it","admin","toilet","sample_collaction"
]

# ---------------- ROOM NAMES ----------------
room_names = {
    "consultation101":"Consultation Room 101",
    "consultation102":"Consultation Room 102",
    "consultation103":"Consultation Room 103",
    "consultation104":"Consultation Room 104",
    "cystoscopy110":"Cystoscopy Room 110",
    "urodynamic109":"Urodynamic Room 109",
    "waiting_area":"Waiting Area",
    "laboratory107":"Laboratory 107",
    "ultrasound106":"Ultrasound Room 106",
    "xray108":"X-Ray Room 108",
    "triage105":"Triage Room 105",
    "staff":"Staff Area",
    "nurse":"Nurse Station",
    "department":"Department",
    "doctor":"Doctor Office",
    "it":"IT Department",
    "admin":"Administration",
    "toilet":"Restroom",
    "Sample_Collaction":"Sample Collection Room"
}

branch_rooms_map = {
    # 🏥 Al Hail
    "alhail": [
        "consultation101",
        "consultation102",
        "consultation103",
        "consultation104",
        "cystoscopy110",
        "urodynamic109",
        "waiting_area",
        "laboratory107",
        "staff",
        "ultrasound106",
        "xray108",
        "triage105",
        "nurse",
        "department",
        "doctor",
        "it",
        "admin",
        "toilet",
        "sample_collaction"
    ],

    # 💊 Pharmacy
    "pharmacy": [
        "pharmacy_area",
        "pharmacy_store",
        "medical_area"
    ],

    # 🏥 Al Amerat (مثال)
    "alamerat": [
        "consultation201",
        "waiting_area_2",
        "lab_201",
        "nurse_station_2"
    ],

    # 🏥 Mabella (مثال)
    "mabella": [
        "consultation301",
        "xray301",
        "ultrasound301",
        "reception"
    ]
}


location_names = {
    "consultation": "Consultation",
    "reception": "Reception",
    "pharmacy": "Pharmacy",
    "lab": "Laboratory",
    "xray": "X-Ray",
    "it": "IT Department"
}
# RECEPTION = [
#     {
#         "name": "welcome",
#         "title": {
#             "en": "How friendly was the reception staff?",
#             "ar": "مدى ترحيب موظف الاستقبال؟"
#         }
#     },
#     {
#         "name": "checkin",
#         "title": {
#             "en": "How easy was the check-in process?",
#             "ar": "سهولة إجراءات التسجيل؟"
#         }
#     },
#     {
#         "name": "waitinfo",
#         "title": {
#             "en": "Was waiting time clearly explained?",
#             "ar": "هل تم توضيح وقت الانتظار؟"
#         }
#     }
# ]

branches = list(branch_rooms_map.keys())

def get_branch_from_location(location):
    for branch, rooms in branch_rooms_map.items():
        if location in rooms:
            return branch
    return "unknown"
# ---------------- HOME ----------------
@app.route('/')
def home():
    return "StarCare System Running 🚀"

# ---------------- FEEDBACK ----------------
# @app.route('/feedback/<location>', methods=['GET', 'POST'])
# def feedback(location):

#     if location not in locations:
#         return "Invalid Location", 404

#     if request.method == 'POST':

#         rating = request.form.get('rating')
#         comment = request.form.get('comment')

#         if not rating:
#             return {"error": "missing rating"}

#         rating = int(rating)

#     # ⭐ 4 و 5 → حفظ مباشر
#         if rating >= 4:

#             collection.insert_one({
#                 "location": location,
#                 "branch": get_branch_from_location(location),  # ✔ مهم
#                 "rating": rating,
#                 "comment": comment,
#                 "phone": None,
#                 "date": datetime.now()
#             })

#             return {"need_phone": False}

#     # ⭐ 1 - 3 → نطلب رقم
#         return {
#             "need_phone": True,
#             "rating": rating,
#             "comment": comment,
#             "location": location
#         }

#     return render_template(
#         "feedback.html",
#         location=location,
#         room_name=room_names.get(location, "Unknown Room"),
#         need_phone=False
#     )
QUESTIONS = {

    "reception": {
        "en": [
            "How welcoming and friendly was the reception team?",
            "How easy was the check-in process?",
            "Were you informed about your waiting time?",
            "How clean and comfortable was the reception area?",
            "Would you recommend our reception service?"
        ],
        "ar": [
            "ما مدى ترحيب موظفي الاستقبال بك؟",
            "ما مدى سهولة إجراءات التسجيل؟",
            "هل تم إبلاغك بمدة الانتظار؟",
            "ما مدى نظافة وراحة منطقة الاستقبال؟",
            "هل توصي بخدمة الاستقبال لدينا؟"
        ]
    },

    "waiting": {
        "en": [
            "How comfortable was the waiting area?",
            "How clean was the waiting area?",
            "How satisfied were you with the waiting time?",
            "How comfortable was the environment?",
            "Would you recommend our waiting area?"
        ],
        "ar": [
            "ما مدى راحة منطقة الانتظار؟",
            "ما مدى نظافة منطقة الانتظار؟",
            "ما مدى رضاك عن مدة الانتظار؟",
            "ما مدى راحة البيئة المحيطة؟",
            "هل توصي بمنطقة الانتظار لدينا؟"
        ]
    },

    "consultation": {
        "en": [
            "Did the doctor listen carefully to your concerns?",
            "Did the doctor explain your treatment clearly?",
            "Was your privacy respected?",
            "How clean was the consultation room?",
            "Would you recommend this doctor?"
        ],
        "ar": [
            "هل استمع الطبيب لمشكلتك باهتمام؟",
            "هل شرح الطبيب العلاج بوضوح؟",
            "هل تم احترام خصوصيتك؟",
            "ما مدى نظافة غرفة العيادة؟",
            "هل توصي بهذا الطبيب؟"
        ]
    },

    "xray": {
        "en": [
            "How comfortable was the X-ray procedure?",
            "Were the instructions clear?",
            "How friendly was the X-ray staff?",
            "How satisfied were you with the appointment speed?",
            "Would you recommend the X-ray department?"
        ],
        "ar": [
            "ما مدى راحتك أثناء الأشعة؟",
            "هل كانت التعليمات واضحة؟",
            "ما مدى تعامل موظفي الأشعة معك؟",
            "ما مدى رضاك عن سرعة الخدمة؟",
            "هل توصي بقسم الأشعة؟"
        ]
    },

    "lab": {
        "en": [
            "How professional was the laboratory staff?",
            "Was the sample collection comfortable?",
            "How clean was the laboratory?",
            "How satisfied were you with the waiting time?",
            "Would you recommend our laboratory?"
        ],
        "ar": [
            "ما مدى احترافية موظفي المختبر؟",
            "هل كانت عملية أخذ العينة مريحة؟",
            "ما مدى نظافة المختبر؟",
            "ما مدى رضاك عن مدة الانتظار؟",
            "هل توصي بالمختبر؟"
        ]
    }, 

    "pharmacy": {
        "en": [
            "How friendly was the pharmacy staff?",
            "Were your medications explained clearly?",
            "How quickly did you receive your medication?",
            "How organized was the pharmacy?",
            "Would you recommend our pharmacy?"
        ],
        "ar": [
            "ما مدى تعامل موظفي الصيدلية؟",
            "هل تم شرح الأدوية بوضوح؟",
            "ما مدى سرعة استلام الدواء؟",
            "ما مدى تنظيم الصيدلية؟",
            "هل توصي بالصيدلية؟"
        ]
    },

    "toilet": {
        "en": [
            "How clean was the restroom?",
            "Were all supplies available?",
            "Did all facilities work properly?",
            "Overall, how satisfied are you with the restroom?"
        ],
        "ar": [
            "ما مدى نظافة دورة المياه؟",
            "هل كانت جميع المستلزمات متوفرة؟",
            "هل كانت جميع المرافق تعمل بشكل صحيح؟",
            "بشكل عام، ما مدى رضاك عن دورة المياه؟"
        ]
    },
    "it": {
        "en": [
            "How quickly did the IT team respond to your request?",
            "Was your issue resolved successfully?",
            "How professional was the IT support staff?",
            "How satisfied are you with the communication during the support process?",
            "Overall, how satisfied are you with the IT service?"
        ],
        "ar": [
            "ما مدى سرعة استجابة فريق تقنية المعلومات لطلبك؟",
            "هل تم حل المشكلة بنجاح؟",
            "ما مدى احترافية فريق تقنية المعلومات؟",
            "ما مدى رضاك عن التواصل أثناء معالجة المشكلة؟",
            "بشكل عام، ما مدى رضاك عن خدمة تقنية المعلومات؟"
        ]
    },
}


# def get_questions(room):
#     room_lower = room.lower()

#     if "reception" in room_lower:
#         return QUESTIONS["reception"]

#     elif "waiting" in room_lower:
#         return QUESTIONS["waiting"]

#     elif "consultation" in room_lower or "doctor" in room_lower:
#         return QUESTIONS["consultation"]

#     elif "xray" in room_lower:
#         return QUESTIONS["xray"]

#     elif "lab" in room_lower:
#         return QUESTIONS["lab"]

#     elif "pharmacy" in room_lower:
#         return QUESTIONS["pharmacy"]

#     elif "toilet" in room_lower:
#         return QUESTIONS["toilet"]

#     return QUESTIONS["consultation"]

def get_questions(location):

    location = location.lower()

    return QUESTIONS.get(location, QUESTIONS["consultation"])



@app.route('/feedback/<branch>/<room>', methods=['GET', 'POST'])
def feedback(branch, room):

    # جلب الأسئلة + اللغة
    lang = request.args.get("lang", "en")

    questions = get_questions(room)

    print(questions)

    room_name = room_names.get(room, room)

    room_lower = room.lower()

# تحديد نوع السؤال من QUESTIONS
    if "reception" in room_lower:
        key = "reception"

    elif "waiting" in room_lower:
        key = "waiting"

    elif "consultation" in room_lower or "doctor" in room_lower:
        key = "consultation"

    elif "xray" in room_lower:
        key = "xray"

    elif "lab" in room_lower:
        key = "lab"

    elif "pharmacy" in room_lower:
        key = "pharmacy"

    elif "toilet" in room_lower:
        key = "toilet"

    else:
        key = "consultation"


    # ================= POST =================

    if request.method == "POST":

        data = request.get_json()

        rating = data.get("rating")
        comment = data.get("comment")

        

        if not rating:
            return jsonify({"error": "missing rating"})

        rating = int(rating)

        if rating >= 4:

            collection.insert_one({
                "branch": branch,
                "location": room,
                "rating": rating,
                "comment": comment,
                "phone": None,
                "name": None,
                "created_at": datetime.now(),

    # ⭐ الجديد
                "answers": {
                    "category": request.get_json().get("category"),
                    "speed": request.get_json().get("speed"),
                    "behavior": request.get_json().get("behavior")
                }
            })

            return jsonify({
                "status": "success",
                "redirect": f"/thankyou/{branch}/{room}"
            })

        # ⭐ low rating
        return jsonify({
            "status": "need_phone",
            "rating": rating,
            "comment": comment
        })

    return render_template(
        "feedback.html",
        room_name=room_name,
        branch=branch,
        room=room,
        location=key,      # ⭐ الجديد
        questions=questions,
        lang=lang
    )


@app.route("/feedback/<branch>/<location>/<room_number>")
def feedback_new(branch, location, room_number):

    lang = request.args.get("lang", "en")

    questions = get_questions(location)

    room_name = f"{location.title()} {room_number}"

    key = location.lower()

    return render_template(
        "feedback.html",
        room_name=room_name,
        branch=branch,
        room=room_number,
        location=key,
        questions=questions,
        lang=lang
    )


from datetime import datetime


@app.route("/save_feedback", methods=["POST"])
def save_feedback():

    data = request.get_json()

    if not data:
        return jsonify({"success": False, "error": "No data"}), 400

    # ⭐ نوع الصفحة (xray / lab / reception ...)
    location = data.get("location")
    room_number = data.get("room_number")
    # ⭐ الإجابات
    answers = data.get("answers", [])

    # ⭐ جلب الأسئلة حسب النوع
    questions_list = QUESTIONS.get(location, {}).get("en", [])

    # ⭐ تحويل الإجابات إلى أسئلة منظمة
    questions = []

    for i in range(len(answers)):
        questions.append({
            "title": questions_list[i] if i < len(questions_list) else f"Q{i+1}",
            "value": answers[i]
        })

    # ⭐ حفظ البيانات
    feedback = {
        "branch": data.get("branch"),

        "location": location,
        "room_number": room_number,

    # هذا الجديد
        "room_display": f"{location.title()} {room_number}",

        "questions": questions,
        "rating": data.get("rating"),

        "comment": data.get("comment", ""),
        "name": data.get("name", "-"),
        "phone": data.get("phone", "-"),

        "created_at": datetime.now(
            ZoneInfo("Asia/Muscat")
        ).strftime("%Y-%m-%d %I:%M %p")
    }

    collection.insert_one(feedback)

    return jsonify({"success": True})


@app.route('/thankyou/<branch>/<room>')
def thankyou(branch, room):

    name = room_names.get(room, room)

    return render_template(
        "thankyou.html",
        branch=branch,
        room=room,
        room_name=name
    )

@app.route("/thankyou/<branch>/<location>/<room_number>")
def thankyou_new(branch, location, room_number):

    room_name = f"{location.title()} {room_number}"

    return render_template(
        "thankyou.html",
        branch=branch,
        location=location,
        room=room_number,
        room_name=room_name
    )
# ---------------- LOGIN ----------------
@app.route('/login', methods=['GET','POST'])
def login():

    if request.method == 'POST':

        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        user = db.users.find_one({"username": username})

        if user and user.get("password") == password:

            session.clear()
            session['admin'] = True
            session['role'] = user.get("role", "user")
            session['username'] = username

            return redirect('/admin')

        return "Wrong credentials"

    return render_template("login.html")

# ---------------- LOGOUT ----------------
@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect('/login')

# ---------------- ADMIN ----------------
@app.route('/admin')
def admin():

    if 'admin' not in session:
        return redirect('/login')

    role = session.get("role")
    username = session.get("username")

    # =====================
    # Filters
    # =====================
    branch = request.args.get("branch")
    location = request.args.get("location")
    room = request.args.get("room")

    # =====================
    # Base query (Branch)
    # =====================
    base_query = {}

    if branch:
        base_query["branch"] = branch

    # =====================
    # Locations (service/department)
    # =====================
    locations = sorted(
        collection.distinct("location", base_query)
    )

    # =====================
    # Rooms (depends on branch + location)
    # =====================
    room_query = dict(base_query)

    if location:
        room_query["location"] = location

    rooms = sorted(
        collection.distinct("room_number", room_query)
    )

    # =====================
    # Main query for table
    # =====================
    query = dict(base_query)

    if location:
        query["location"] = location

    if room:
        query["room_number"] = room   # ✅ مهم جداً (صححناه)

    # =====================
    # Permissions
    # =====================
    user = db.users.find_one({"username": username})
    allowed_locations = user.get("locations", []) if user else []

    if role != "admin":
        query["location"] = {"$in": allowed_locations}

    # =====================
    # DATA
    # =====================
    data = list(collection.find(query).sort("created_at", -1))

    for i in data:
        i["facility"] = int(i.get("facility") or 0)
        i["it"] = int(i.get("it") or 0)
        i["medical"] = int(i.get("medical") or 0)
        i["nursing"] = int(i.get("nursing") or 0)
        i["other"] = int(i.get("other") or 0)

        i["rating"] = safe_float(i.get("rating"))

    # =====================
    # STATS
    # =====================
    total = len(data)

    valid_ratings = [i["rating"] for i in data if i["rating"] > 0]
    avg = round(sum(valid_ratings) / len(valid_ratings), 2) if valid_ratings else 0

    stats = {}

    for i in data:
        loc = i.get("location")

        if loc not in stats:
            stats[loc] = {"count": 0, "total": 0}

        stats[loc]["count"] += 1

        r = i.get("rating")
        if isinstance(r, (int, float)):
            stats[loc]["total"] += r

    stats_list = [
        {
            "location": l,
            "count": v["count"],
            "avg": round(v["total"] / v["count"], 2) if v["count"] else 0
        }
        for l, v in stats.items()
    ]

    # =====================
    # BRANCHES
    # =====================
    branches = list(branch_rooms_map.keys())

    return render_template(
        "dashboard.html",
        data=data,
        total_feedback=total,
        avg_rating=avg,

        branches=branches,
        locations=locations,
        rooms=rooms,

        stats=stats_list,

        role=role,

        active_branch=branch,
        active_location=location,
        active_room=room
    )
#اضافه مستخدمين
@app.route('/add_user', methods=['GET', 'POST'])
def add_user():

    if request.method == 'POST':

        username = request.form.get("username")
        password = request.form.get("password")
        role = request.form.get("role")

        # 🔥 مهم: lists
        branches = request.form.getlist("branches")
        locations = request.form.getlist("locations")

        db.users.insert_one({
            "username": username,
            "password": password,
            "role": role,

            # 🔥 صلاحيات
            "branches": branches,
            "locations": locations
        })

        return redirect("/manage_users")

    return render_template("add_user.html")

#حذف مستخدمين 
@app.route('/delete_user/<username>')
def delete_user(username):

    if session.get("role") != "admin":
        return "Not Allowed", 403

    db.users.delete_one({"username": username})

    return redirect('/manage_users')

    #عرض مستخدمين 
@app.route('/manage_users')
def manage_users():

    if session.get("role") != "admin":
        return "Not Allowed", 403

    users = list(db.users.find())

    return render_template("manage_users.html", users=users)

@app.route('/add_role', methods=['GET', 'POST'])
def add_role():

    if session.get("role") != "admin":
        return "Not Allowed", 403

    if request.method == 'POST':

        role = request.form['role']
        locations = request.form.getlist('locations')

        db.roles.insert_one({
            "role": role,
            "locations": locations
        })

        return redirect('/admin')

    return render_template("add_role.html")

@app.route('/api/feedback')
def api_feedback():
    data = list(collection.find().sort("date", -1))

    for i in data:
        i["_id"] = str(i["_id"])
        i["date"] = str(i["date"])

    return {"data": data}

@app.route('/delete_feedback/<id>')
def delete_feedback(id):

    if 'admin' not in session:
        return redirect('/login')

    collection.delete_one({
        "_id": ObjectId(id)
    })

    return redirect('/admin')
# ---------------- PDF ----------------
@app.route('/download_pdf')
def download_pdf():

    if 'admin' not in session:
        return redirect('/login')

    # 👇 هنا تحطين الكود الجديد مباشرة
    query = {}

    branch = request.args.get("branch") or session.get("active_branch")
    location = request.args.get("location")
    role = session.get("role")
    category = request.args.get("category")

    filters = []

# 🔹 branch filter
    if branch and branch != "all":
        filters.append({"branch": branch})

# 🔹 role filter (IT)
    if role != "admin":
        user = db.users.find_one({"username": session.get("username")})
        allowed_locations = user.get("locations", [])

        if allowed_locations:
            filters.append({"location": {"$in": allowed_locations}})

# 🔹 dropdown location
    if location:
        filters.append({"location": location})

    if category:
        filters.append({"category": category})

# 🔥 combine correctly
    if filters:
        query = {"$and": filters}
    else:
        query = {}
        
    # 📦 جلب البيانات
    data = list(collection.find(query))

    for i in data:

        i["category"] = i.get("category", "-")
        i["branch"] = i.get("branch", "-")
        i["location"] = i.get("location", "-")
        i["room_number"] = i.get("room_number", "-")
        i["comment"] = i.get("comment", "-")
        i["questions"] = i.get("questions", [])

    # حساب Overall Rating
        if i.get("rating") is not None:
            i["rating"] = float(i["rating"])
        else:
            values = []

            for q in i["questions"]:
                try:
                    values.append(float(q.get("value", 0)))
                except:
                    pass

            i["rating"] = round(sum(values) / len(values), 1) if values else 0

    # (حذفنا هذا الغلط القديم)
    # data = list(collection.find()) ❌

    # 🔢 الحسابات
    total = len(data)

    five_star = len([x for x in data if x.get("rating", 0) == 5])
    four_star = len([x for x in data if x.get("rating", 0) == 4])
    three_star = len([x for x in data if x.get("rating", 0) == 3])
    two_star = len([x for x in data if x.get("rating", 0) == 2])
    one_star = len([x for x in data if x.get("rating", 0) == 1])

    avg = round(
        sum(i["rating"] for i in data) / total,
        2
    ) if total else 0

    stats = {}

    for item in data:

        loc = item.get("location", "-")

        overall = item.get("rating")

        if overall is None:

            questions = item.get("questions", [])

            values = []

            for q in questions:
                try:
                    values.append(float(q.get("value", 0)))
                except:
                    pass

            overall = round(sum(values) / len(values), 2) if values else 0

        if loc not in stats:
            stats[loc] = {
                "count": 0,
                "total": 0
            }

        stats[loc]["count"] += 1
        stats[loc]["total"] += overall



    # 📄 PDF generation
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)

    # pdfmetrics.registerFont(TTFont('Helvetica', 'Helvetica'))

    styles = getSampleStyleSheet()
    # styles['Normal'].fontName = 'Arabic'
    # styles['Title'].fontName = 'Arabic'
    # styles['Heading2'].fontName = 'Arabic'
    elements = []

    logo_path = os.path.join(app.root_path, "static", "logowhite.jpeg")

    logo = Image(logo_path)
    logo.drawHeight = 100
    logo.drawWidth = 180

    elements.append(logo)
    elements.append(Spacer(1, 15))

    # ===========================
# Title
# ===========================

    elements.append(
        Paragraph(
            "<font size='22' color='#00A79B'><b>StarCare Hospital Feedback Report</b></font>",
            styles["Title"]
        )
    )

    elements.append(Spacer(1, 8))

    elements.append(
        HRFlowable(
            width="100%",
            thickness=2,
            color=colors.HexColor("#00A79B")
        )
    )

    elements.append(Spacer(1, 15))

# ===========================
# Report Information
# ===========================
    oman_time = datetime.now(ZoneInfo("Asia/Muscat"))
    report_info = [

        ["Report Date", oman_time.strftime("%d %b %Y")],

        ["Report Time", oman_time.strftime("%I:%M %p")],

        ["Branch", branch if branch else "All Branches"],

        ["Location", location if location else "All Locations"],

        ["Category", category if category else "All Categories"]

    ]

    info_table = Table(
        report_info,
        colWidths=[130, 260]
    )

    info_table.setStyle(TableStyle([

        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor("#00A79B")),

        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),

        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),

        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),

    ]))

    elements.append(info_table)

    elements.append(Spacer(1, 18))

# ===========================
# Summary
# ===========================

    summary_table = Table([

        ["Total Feedback", total],

        ["Average Rating", f"{avg} ⭐"],

        ["5 Stars", five_star],

        ["4 Stars", four_star],

        ["3 Stars", three_star],

        ["2 Stars", two_star],

        ["1 Star", one_star],

    ], colWidths=[180, 120])

    summary_table.setStyle(TableStyle([

        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor("#EAF9F8")),

        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

        ('BOX', (0, 0), (-1, -1), 1, colors.black),

        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),

    ]))

    elements.append(summary_table)

    elements.append(Spacer(1, 18))
    

    if avg >= 4.5:
        summary = "Overall patient satisfaction is excellent."
    elif avg >= 4:
        summary = "Overall patient satisfaction is good."
    elif avg >= 3:
        summary = "Patient satisfaction needs improvement."
    else:
        summary = "Immediate action is recommended."

    elements.append(
        Paragraph(f"<b>Executive Summary:</b> {summary}", styles['Normal'])
    )

    if branch:
        elements.append(
            Paragraph(f"Branch: {branch.upper()}", styles['Heading2'])
        )
    else:
        elements.append(
            Paragraph("Branch: ALL BRANCHES", styles['Heading2'])
        )

    elements.append(Spacer(1, 20))

    elements.append(
    Paragraph("<b>Department Performance Summary</b>", styles["Heading2"])
)

    elements.append(Spacer(1,10))

    rows = [["Department", "Feedback", "Average", "Status"]]

    for loc, v in stats.items():

        avg_loc = round(v["total"] / v["count"], 2)

        if avg_loc >= 4.5:
            status = "Excellent"

        elif avg_loc >= 4:
            status = "Good"

        elif avg_loc >= 3:
            status = "Needs Improvement"

        else:
            status = "Critical"

        rows.append([
            location_names.get(loc, loc),
            str(v["count"]),
            f"{avg_loc} ⭐",
            status
        ])

    table = Table(
        rows,
        colWidths=[180,70,70,120]
    )

    table.setStyle(TableStyle([

    ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#00A79B")),

    ('TEXTCOLOR',(0,0),(-1,0),colors.white),

    ('FONTNAME',(0,0),(-1,0),"Helvetica-Bold"),

    ('ALIGN',(0,0),(-1,-1),"CENTER"),

    ('GRID',(0,0),(-1,-1),0.5,colors.grey),

    ('ROWBACKGROUNDS',(0,1),(-1,-1),[
        colors.white,
        colors.HexColor("#F7FCFC")
    ]),

    ('BOTTOMPADDING',(0,0),(-1,-1),8),

    ]))

    elements.append(table)

    elements.append(Spacer(1,25))

    elements.append(
        Paragraph("Detailed Feedback", styles['Heading2'])
    )

    details = [[
        "Date",
        "Time",
        "Branch",
        "Department",
        "Overall",
        "Questions",
        "Comment",
        "Name",
        "Phone"
    ]]

    for item in data:

    # -----------------------
    # Date & Time
    # -----------------------
        date_obj = item.get("created_at") or item.get("date")

        date_str = "-"
        time_str = "-"

        if isinstance(date_obj, datetime):

            if date_obj.tzinfo is None:
                date_obj = date_obj.replace(tzinfo=ZoneInfo("UTC"))

            date_obj = date_obj.astimezone(ZoneInfo("Asia/Muscat"))

            date_str = date_obj.strftime("%Y-%m-%d")
            time_str = date_obj.strftime("%I:%M %p")

    # -----------------------
    # Overall Rating
    # -----------------------
        overall = item.get("rating")

        if overall is None:

            values = []

            for q in item.get("questions", []):
                try:
                    values.append(float(q.get("value", 0)))
                except:
                    pass

            overall = round(sum(values) / len(values), 1) if values else 0

        stars = f"{overall}/5"

    # -----------------------
    # Questions
    # -----------------------
        questions_text = ""

        for q in item.get("questions", []):

            questions_text += (
                f"• {q.get('title')}: "
                f"{q.get('value')}/5<br/>"
            )

        if not questions_text:
            questions_text = "-"

        comment = item.get("comment") or "-"

        details.append([
            date_str,
            time_str,
            item.get("branch", "-"),
            item.get("location", "-"),
            stars,
            Paragraph(questions_text, styles["BodyText"]),
            comment,
            item.get("name", "-"),
            item.get("phone", "-")
        ])

    details_table = Table(
        details,
        colWidths=[
            50,   # Date
            45,   # Time
            55,   # Branch
            60,   # Department
            45,   # Overall
            210,  # Questions
            90,   # Comment
            70,   # Name
            70    # Phone
        ],
        repeatRows=1
    )


    details_table.setStyle(TableStyle([

        ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#00A79B")),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),"Helvetica-Bold"),
        ('FONTSIZE',(0,0),(-1,0),10),

        ('FONTSIZE',(0,1),(-1,-1),8),

        ('GRID',(0,0),(-1,-1),0.6,colors.black),
        ('BOX',(0,0),(-1,-1),1,colors.black),

        ('ALIGN',(0,0),(-1,0),"CENTER"),
        ('ALIGN',(0,1),(4,-1),"CENTER"),
        ('VALIGN',(0,0),(-1,-1),"TOP"),

        ('LEFTPADDING',(0,0),(-1,-1),5),
        ('RIGHTPADDING',(0,0),(-1,-1),5),
        ('TOPPADDING',(0,0),(-1,-1),6),
        ('BOTTOMPADDING',(0,0),(-1,-1),6),

        ('ROWBACKGROUNDS',(0,1),(-1,-1),[
            colors.white,
            colors.HexColor("#F5F9F9")
        ]),

    ]))

    elements.append(
        Paragraph(
            "<b>Patient Feedback Details</b>",
            styles["Heading2"]
        )
    )

    elements.append(Spacer(1,8))
    elements.append(details_table)

    elements.append(Spacer(1, 20))
    elements.append(
        Paragraph("Low Rating Cases", styles['Heading2'])
    )

    # elements.append(
    #     Paragraph(
    #         "<b>Critical Feedback Cases</b>",
    #         styles["Heading2"]
    #     )
    # )

    elements.append(Spacer(1,8))

    low_rows = [[
        "Location",
        "Rating",
        "Comment",
        "Phone"
    ]]

    for item in data:

        if item.get("rating",0) <= 3:

            stars = "⭐"*int(item.get("rating",0))

            comment = item.get("comment","-")

            if len(comment) > 40:
                comment = comment[:40] + "..."

            low_rows.append([
                room_names.get(item.get("location"), item.get("location")),
                stars,
                comment,
                item.get("phone","-")
            ])

    if len(low_rows) > 1:

        low_table = Table(
            low_rows,
            colWidths=[140,70,220,90]
        )

        low_table.setStyle(TableStyle([

            ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#E74C3C")),

            ('TEXTCOLOR',(0,0),(-1,0),colors.white),

            ('FONTNAME',(0,0),(-1,0),"Helvetica-Bold"),

            ('GRID',(0,0),(-1,-1),0.5,colors.grey),

            ('ROWBACKGROUNDS',(0,1),(-1,-1),[
                colors.white,
                colors.HexColor("#FFF5F5")
            ]),

            ('BOTTOMPADDING',(0,0),(-1,-1),8),

        ]))

        elements.append(low_table)

    else:

        elements.append(
            Paragraph(
                "No critical feedback found.",
                styles["Normal"]
            )
        )

    elements.append(Spacer(1,20))


    

    elements.append(Spacer(1, 20))
    elements.append(
        Paragraph("Management Recommendations", styles['Heading2'])
    )

    if avg >= 4.5:
        recommendation = """
        • Maintain current service quality.
        • Continue monitoring patient satisfaction.
        """
    elif avg >= 4:
        recommendation = """
        • Improve patient communication.
        • Monitor waiting times.
        """
    elif avg >= 3:
        recommendation = """
        • Review patient complaints.
        • Improve waiting time and service delivery.
        """
    else:
        recommendation = """
        • Immediate management review required.
        • Contact dissatisfied patients.
        • Investigate recurring issues.
        """

    elements.append(
        Paragraph(recommendation.replace("\n", "<br/>"), styles['BodyText'])
    )

    elements.append(Spacer(1,25))

    elements.append(
        HRFlowable(
            width="100%",
            thickness=1,
            color=colors.grey
        )
    )

    elements.append(Spacer(1,8))

    elements.append(
        Paragraph(
            "<font size='9' color='grey'>"
            "Generated automatically by StarCare Hospital Feedback System"
            "<br/>Confidential Report"
            "</font>",
            styles["Normal"]
        )
    )

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"starcare_{branch if branch else 'all_branches'}_report.pdf"

    return Response(
        pdf,
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename={filename}'
        }
    )




@app.route('/download_excel')
def download_excel():

    if 'admin' not in session:
        return redirect('/login')

    query = {}

    branch = request.args.get("branch") or session.get("active_branch")
    location = request.args.get("location")
    room = request.args.get("room")
    role = session.get("role")

    filters = []

    # =====================
    # Branch filter
    # =====================
    if branch and branch != "all":
        filters.append({"branch": branch})

    # =====================
    # Permissions
    # =====================
    if role != "admin":
        user = db.users.find_one({"username": session.get("username")})
        allowed_locations = user.get("locations", [])
        if allowed_locations:
            filters.append({"location": {"$in": allowed_locations}})

    # =====================
    # Location filter (department)
    # =====================
    if location:
        filters.append({"location": location})

    # =====================
    # Room filter (IMPORTANT)
    # =====================
    if room:
        filters.append({"room_number": room})

    if filters:
        query = {"$and": filters}

    data = list(collection.find(query))


    # if data:
    #     print(data[0]["date"])
    #     print(type(data[0]["date"]))
    # ===========================
    # Excel file
    # ===========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Report"

    # ===========================
    # Title
    # ===========================
    ws["A1"] = "StarCare Hospital Feedback Report"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="00A79B")

    oman_now = datetime.now(ZoneInfo("Asia/Muscat"))

    ws["A3"] = "Report Date"
    ws["B3"] = oman_now.strftime("%Y-%m-%d")

    ws["A4"] = "Report Time"
    ws["B4"] = oman_now.strftime("%I:%M %p")

    ws["A5"] = "Branch"
    ws["B5"] = branch if branch else "All Branches"

    ws["A6"] = "Location"
    ws["B6"] = location if location else "All Departments"

    ws["A7"] = "Room"
    ws["B7"] = room if room else "All Rooms"

    # ===========================
    # Statistics
    # ===========================
    ratings = []

    for item in data:

        if item.get("rating") is not None:
            ratings.append(float(item["rating"]))

        else:
            qs = item.get("questions", [])

            vals = []

            for q in qs:
                try:
                    vals.append(float(q.get("value", 0)))
                except:
                    pass

            if vals:
                ratings.append(round(sum(vals) / len(vals), 2))
            else:
                ratings.append(0)

    total = len(ratings)
    avg = round(sum(ratings)/total, 2) if total else 0

    five = len([r for r in ratings if r == 5])
    four = len([r for r in ratings if r == 4])
    three = len([r for r in ratings if r == 3])
    two = len([r for r in ratings if r == 2])
    one = len([r for r in ratings if r == 1])

    ws["D3"] = "Total Feedback"
    ws["E3"] = total

    ws["D4"] = "Average Rating"
    ws["E4"] = avg

    ws["D5"] = "5 Stars"
    ws["E5"] = five

    ws["D6"] = "4 Stars"
    ws["E6"] = four

    ws["D7"] = "3 Stars"
    ws["E7"] = three

    ws["D8"] = "2 Stars"
    ws["E8"] = two

    ws["D9"] = "1 Star"
    ws["E9"] = one

    # ===========================
    # Table Header
    # ===========================
    row = 12

    headers = [
        "Date",
        "Time",
        "Branch",
        "Department",
        "Room",
        "Overall Rating",
        "Questions & Answers",
        "Comment",
        "Name",
        "Phone"
    ]

    for col, head in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = head
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="00A79B")
        cell.alignment = Alignment(horizontal="center")

    # ===========================
    # Data rows
    # ===========================
    for item in data:

        date_obj = item.get("created_at")

        if not date_obj:
            date_obj = item.get("date")

    # إذا كان String نحوله إلى datetime
        if isinstance(date_obj, str):
            try:
                date_obj = datetime.strptime(
                    date_obj,
                    "%Y-%m-%d %I:%M %p"
                )
                date_obj = date_obj.replace(tzinfo=ZoneInfo("Asia/Muscat"))
            except:
                date_obj = None

    # إذا كان datetime نحوله لتوقيت عمان
        elif isinstance(date_obj, datetime):

        # إذا datetime بدون timezone
            if date_obj.tzinfo is None:
                date_obj = date_obj.replace(tzinfo=ZoneInfo("UTC"))

            date_obj = date_obj.astimezone(ZoneInfo("Asia/Muscat"))

    # تنسيق العرض
        if date_obj:
            date_str = date_obj.strftime("%Y-%m-%d")
            time_str = date_obj.strftime("%I:%M %p")
        else:
            date_str = "-"
            time_str = "-"
            
        questions = item.get("questions", [])

        qa = ""

        for i, q in enumerate(questions, start=1):
            qa += f"{i}. {q.get('title')}\n"
            qa += f"   Rating: {q.get('value')}/5\n\n"

        rating = item.get("rating")

        if rating is None:
            qs = item.get("questions", [])
            vals = []

            for q in qs:
                try:
                    vals.append(float(q.get("value", 0)))
                except:
                    pass

            rating = round(sum(vals) / len(vals), 1) if vals else 0

        ws.append([
            date_str,
            time_str,
            item.get("branch", "-"),
            item.get("location", "-"),
            item.get("room_number", "-"),
            f"{rating}/5",
            qa,
            item.get("comment", ""),
            item.get("name", "-"),
            item.get("phone", "-"),
        ])

    # ===========================
    # Styling
    # ===========================
    thin = Side(style="thin")

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(length + 5, 40)

    ws.freeze_panes = "A13"

    # ===========================
    # Return file
    # ===========================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=feedback_report.xlsx"
        }
    )


from flask import request, render_template, redirect, session
from datetime import datetime
from collections import Counter

@app.route('/question_analytics')
def question_analytics():

    if 'admin' not in session:
        return redirect('/login')

    # ================= GET FILTERS =================
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    branch = request.args.get("branch")
    location = request.args.get("location")

    data = list(collection.find())

    filtered = []

    for i in data:

        # ================= DATE FIX =================
        db_date = i.get("date")

        if isinstance(db_date, str):
            try:
                db_date = datetime.strptime(db_date, "%Y-%m-%d")
            except:
                db_date = None

        i["date"] = db_date

        # ================= FILTER CHECK =================
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            if not db_date or db_date < start:
                continue

        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            if not db_date or db_date > end:
                continue

        if branch and i.get("branch") != branch:
            continue

        if location and i.get("location") != location:
            continue

        filtered.append(i)

    data = filtered

    # ================= STATS =================
    ratings = [i.get("rating", 0) for i in data if isinstance(i.get("rating"), (int, float))]
    avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else 0

    category_count = Counter(i.get("category", "unknown") for i in data)
    speed_count = Counter(i.get("speed", "unknown") for i in data)
    behavior_count = Counter(i.get("behavior", "unknown") for i in data)

    excellent_count = len([i for i in data if i.get("behavior") == "excellent"])
    bad_count = len([i for i in data if i.get("behavior") == "bad"])

    # ================= ROOMS =================
    room_stats = {}

    for i in data:
        room = i.get("location", "unknown")
        rating = i.get("rating", 0)

        room_stats.setdefault(room, {"count": 0, "total": 0})
        room_stats[room]["count"] += 1

        if isinstance(rating, (int, float)):
            room_stats[room]["total"] += rating

    room_labels = list(room_stats.keys())
    room_counts = [v["count"] for v in room_stats.values()]
    room_avgs = [round(v["total"] / v["count"], 2) if v["count"] else 0 for v in room_stats.values()]

    # ================= FILTER OPTIONS =================
    branches = sorted(set(i.get("branch") for i in collection.find() if i.get("branch")))
    rooms = sorted(set(i.get("location") for i in collection.find() if i.get("location")))

    return render_template(
        "question_analytics.html",
        data=data,
        avg_rating=avg_rating,
        category_count=category_count,
        speed_count=speed_count,
        behavior_count=behavior_count,
        excellent_count=excellent_count,
        bad_count=bad_count,
        room_labels=room_labels,
        room_counts=room_counts,
        room_avgs=room_avgs,
        branches=branches,
        rooms=rooms,
        selected_branch=branch,
        selected_location=location,
        start_date=start_date,
        end_date=end_date
    )
# ---------------- ANALYTICS ----------------
@app.route('/analytics')
def analytics():

    if 'admin' not in session:
        return redirect('/login')

    branch = request.args.get("branch")
    location = request.args.get("location")
    room = request.args.get("room")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = {}

    filters = []

    if branch:
        filters.append({"branch": branch})

    if location:
        filters.append({"location": location})

    if room:
        filters.append({"room_number": room})

    if filters:
        query["$and"] = filters

    query["$or"] = [
        {"status": {"$exists": False}},
        {"status": {"$ne": "resolved"}}
    ]

    data = list(collection.find(query))


    filtered = []

    for item in data:

        date_obj = item.get("created_at") or item.get("date")

        if isinstance(date_obj, str):
            try:
                date_obj = datetime.strptime(date_obj, "%Y-%m-%d %I:%M %p")
            except:
                try:
                    date_obj = datetime.strptime(date_obj, "%Y-%m-%d")
                except:
                    date_obj = None

        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            if not date_obj or date_obj < start:
                continue

        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            if not date_obj or date_obj > end:
                continue

        filtered.append(item)

    data = filtered

    # 📈 stats
    stats = {}

    for i in data:

        loc = i.get("location", "Unknown")

    # calculate rating from questions
        if i.get("rating"):
            rating = float(i.get("rating"))

        elif i.get("questions"):

            values = [
                q.get("value",0)
                for q in i["questions"]
            ]

            rating = sum(values) / len(values)

        else:
            rating = 0

        if loc not in stats:
            stats[loc] = {
            "count": 0,
            "total": 0,
            "ids": []
        }

        stats[loc]["count"] += 1
        stats[loc]["total"] += rating
        stats[loc]["ids"].append(str(i["_id"]))

    chart_data = [
        {
            "ids": v["ids"],
            "location": room_names.get(loc, loc),
            "count": v["count"],
            "avg": round(v["total"] / v["count"], 2)
        }
        for loc, v in stats.items()
    ]

    # 🚨 Attention rooms
    attention_rooms = []

    for item in chart_data:

        if item["avg"] <= 3:

            attention_rooms.append(item)

    # Create attention alerts
    for loc, v in stats.items():

        avg = round(v["total"] / v["count"], 2)

        if avg <= 3:

            existing = alerts_collection.find_one({
                "location": loc,
                "status": "open"
            })

            if not existing:

                alerts_collection.insert_one({
                    "location": loc,
                    "average_rating": avg,
                    "status": "open",
                    "created_at": datetime.now()
                })

    branches = sorted(collection.distinct("branch"))
    locations = sorted(collection.distinct("location"))
    rooms = sorted(collection.distinct("room_number"))

    alerts = list(alerts_collection.find({
        "status":"open"
    }))
    return render_template(
        "analytics.html",
        data=chart_data,
        attention_rooms=attention_rooms,
        branches=branches,
        locations=locations,
        rooms=rooms,
        alerts=alerts,
        selected_branch=branch,
        selected_location=location,
        selected_room=room,
        start_date=start_date,
        end_date=end_date
    )



@app.route('/resolve_alert/<id>')
def resolve_alert(id):

    if 'admin' not in session:
        return redirect('/login')

    collection.update_one(
        {"_id": ObjectId(id)},
        {
            "$set":{
                "status":"resolved"
            }
        }
    )

    return redirect('/analytics')
# ---------------- GENERATE QR ----------------


@app.route('/qr_generator')
def qr_generator():

    if 'admin' not in session:
        return redirect('/login')

    return render_template("qr_generator.html", qr_image=None, url=None)


@app.route('/generate_qr', methods=['GET', 'POST'])
def generate_qr():

    if 'admin' not in session:
        return redirect('/login')

    qr_image = None
    url = None

    if request.method == "POST":

        branch = request.form.get("branch")

        location = request.form.get("location").strip().lower()

        room_number = request.form.get("room_number").strip()

        url = f"https://starcare-feedback-1.onrender.com/feedback/{branch}/{location}/{room_number}"
        qr = qrcode.make(url)

        buffer = BytesIO()
        qr.save(buffer, format="PNG")
        buffer.seek(0)

        img_base64 = base64.b64encode(buffer.getvalue()).decode()

        existing = db.qr_codes.find_one({
            "branch": branch,
            "location": location,
            "room_number": room_number
        })

        if not existing:
            db.qr_codes.insert_one({

                "branch": branch,

                "location": location,

                "room_number": room_number,

                "room_display": f"{location.title()} {room_number}",

                "url": url,

                "qr_image": img_base64,

                "created_at": datetime.now()

            })

        qr_image = img_base64

    return render_template(
        "qr_generator.html",
        qr_image=qr_image,
        url=url
    )
# ---------------- QR DASHBOARD ----------------
@app.route('/qr_dashboard')
def qr_dashboard():

    if 'admin' not in session:
        return redirect('/login')

    user = db.users.find_one({"username": session["username"]})

    if user.get("role") != "admin":
        return "Not allowed"

    rooms_data = []

    # 1️⃣ الباركودات القديمة (من map)
    for branch, rooms in branch_rooms_map.items():

        for room in rooms:

            url = f"{BASE_URL}/feedback/{branch}/{room}"

            rooms_data.append({
                "branch": branch,
                "room": room,
                "name": room_names.get(room, room),

                "count": collection.count_documents({
                    "branch": branch,
                    "location": room
                }),

                "qr": url
            })

    # 2️⃣ الباركودات الجديدة (من DB)
    # 2️⃣ الباركودات الجديدة (من DB)
    qrs = list(db.qr_codes.find())

    for qr in qrs:

        feedback_count = collection.count_documents({
            "branch": qr.get("branch"),
            "location": qr.get("location"),
            "room_number": qr.get("room_number")
        })

        rooms_data.append({

            "_id": str(qr["_id"]),

            "branch": qr.get("branch"),

            "location": qr.get("location"),

            "room_number": qr.get("room_number"),

            "room_display": qr.get(
                "room_display",
                f'{qr.get("location").title()} {qr.get("room_number")}'
            ),

            "count": feedback_count,

            "qr": qr.get("url"),

            "qr_image": qr.get("qr_image")
        })

    return render_template("qr_dashboard.html", rooms=rooms_data)


@app.route("/delete_qr/<qr_id>")
def delete_qr(qr_id):

    if "admin" not in session:
        return redirect("/login")

    try:
        db.qr_codes.delete_one({
            "_id": ObjectId(qr_id)
        })
    except:
        return "Invalid ID"

    return redirect("/qr_dashboard")
# @app.route('/fix_branch')
# def fix_branch():

#     for i in collection.find():

#         if "branch" not in i or not i.get("branch"):

#             collection.update_one(
#                 {"_id": ObjectId(i["_id"])},
#                 {"$set": {"branch": get_branch_from_location(i["location"])}}
#             )

#     return "DONE"


# ---------------- RUN ----------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)